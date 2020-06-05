import shutil
import os

import openpyxl

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from datetime import datetime
import time

path = "C:\\Users\\Usuário\\Desktop"
diaNegocios = "04/06/2020"
file = 'diariamente.xlsx'
planilha01 = 'Atualizadas'
planilha02 = 'Desatualizadas'

original = r'' + str(file)
alvo = r'' + str(path) + '\\' + str(file)

shutil.copyfile(original, alvo)

os.chdir(path)

ultimoDia = diaNegocios.replace('/', '_')
os.rename(r'diariamente.xlsx', r'' + str(ultimoDia) + '.xlsx')

file = str(ultimoDia) + '.xlsx'

listAtuais = []
listAntigo = []
listStocks = []

wb = openpyxl.load_workbook(str(file))
sheet = wb[str(planilha01)]

for i in range(2, 350, 1):
    nameStock = sheet.cell(row=i, column=1).value
    listStocks.append(nameStock)

total = 0

j = 1

totalDesa = 1

fund = webdriver.Chrome(ChromeDriverManager().install())
fund.get("http://www.fundamentus.com.br/detalhes.php?papel=")
fund.maximize_window()

for i in listStocks:
    caixa = fund.find_element_by_id('completar')
    caixa.send_keys(i)
    login_attempt = fund.find_element_by_xpath('/html/body/div[1]/div[1]/form/fieldset/input[2]')
    login_attempt.submit()
    time.sleep(1.5)
    try:
        data = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[2]/td[4]').text

    except:
        print('Opa tivemos um erro aqui com a ação ' + i)

    else:

        if data == diaNegocios:
            listAtuais.append(i)

            vard = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[2]/td[2]').text
            vard_n = vard.replace(',', '.')

            varm = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[3]/td[2]/span/font').text
            varm_n = varm.replace(',', '.')

            var30d = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[4]/td[2]/span/font').text
            var30d_n = var30d.replace(',', '.')

            var12m = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[5]/td[2]/span/font').text
            var12m_n = list(var12m)
            for x, w in zip(var12m_n, range(0, len(var12m_n), 1)):
                if x == '.':
                    del var12m_n[w]
                elif x == ',':
                    var12m_n[w] = '.'
            var12m_n = ''.join(var12m_n)
            var12m_n = str(var12m_n)

            var2020 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[6]/td[2]/span/font').text
            var2020_n = list(var2020)
            for x, w in zip(var2020_n, range(0, len(var2020_n), 1)):
                if x == '.':
                    del var2020_n[w]
                elif x == ',':
                    var2020_n[w] = '.'
            var2020_n = ''.join(var2020_n)
            var2020_n = str(var2020_n)

            var2019 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[7]/td[2]/span/font').text
            var2019_n = list(var2019)
            for x, w in zip(var2019_n, range(0, len(var2019_n), 1)):
                if x == '.':
                    del var2019_n[w]
                elif x == ',':
                    var2019_n[w] = '.'
            var2019_n = ''.join(var2019_n)
            var2019_n = str(var2019_n)

            var2018 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[8]/td[2]/span/font').text
            var2018_n = list(var2018)
            for x, w in zip(var2018_n, range(0, len(var2018_n), 1)):
                if x == '.':
                    del var2018_n[w]
                elif x == ',':
                    var2018_n[w] = '.'
            var2018_n = ''.join(var2018_n)
            var2018_n = str(var2018_n)

            var2017 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[9]/td[2]/span/font').text
            var2017_n = list(var2017)
            for x, w in zip(var2017_n, range(0, len(var2017_n), 1)):
                if x == '.':
                    del var2017_n[w]
                elif x == ',':
                    var2017_n[w] = '.'
            var2017_n = ''.join(var2017_n)
            var2017_n = str(var2017_n)

            var2016 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[10]/td[2]/span/font').text
            var2016_n = list(var2016)
            for x, w in zip(var2016_n, range(0, len(var2016_n), 1)):
                if x == '.':
                    del var2016_n[w]
                elif x == ',':
                    var2016_n[w] = '.'
            var2016_n = ''.join(var2016_n)
            var2016_n = str(var2016_n)

            var2015 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[11]/td[2]/span/font').text
            var2015_n = list(var2015)
            for x, w in zip(var2015_n, range(0, len(var2015_n), 1)):
                if x == '.':
                    del var2015_n[w]
                elif x == ',':
                    var2015_n[w] = '.'
            var2015_n = ''.join(var2015_n)
            var2015_n = str(var2015_n)

            vpa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[3]/td[6]/span').text
            vpa_n = vpa.replace(',', '.')

            roe = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[9]/td[6]/span').text
            roe_n = roe.replace(',', '.')

            pa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[1]/td[4]').text
            pa_n = pa.replace(',', '.')

            p_vpa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[3]/td[4]').text
            p_vpa_n = p_vpa.replace(',', '.')

            pl = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[2]/td[4]').text
            pl_n = pl.replace(',', '.')

            lpa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[2]/td[6]').text
            lpa_n = lpa.replace(',', '.')

            divbppa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[11]/td[6]').text
            divbppa_n = divbppa.replace(',', '.')

            setor = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[4]/td[2]').text

            empresa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[3]/td[2]').text

            balanco = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[2]/tbody/tr[1]/td[4]').text

            sheet = wb[str(planilha01)]

            listDados = [i, empresa, setor, roe_n, vpa_n, pa_n, p_vpa_n, pl_n, lpa_n,
                         divbppa_n, vard_n, varm_n, var30d_n, var12m_n, var2020_n, var2019_n,
                         var2018_n, var2017_n, var2016_n, var2015_n, balanco, data]

            j = j + 1

            for coluna in range(0, len(listDados), 1):
                sheet.cell(row=j, column=coluna + 1).value = listDados[coluna]
                data_e_hora_atuais = datetime.now()
                data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')
                sheet.cell(row=j, column=len(listDados) + 1).value = data_e_hora_em_texto

            wb.save(str(file))

        elif data != 'diaNegocios':
            listAntigo.append(i)

            sheet = wb[str(planilha02)]

            listDados = [i, data]

            totalDesa = totalDesa + 1

            for coluna in range(0, len(listDados), 1):
                sheet.cell(row=totalDesa, column=coluna + 1).value = listDados[coluna]
                data_e_hora_atuais = datetime.now()
                data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')
                sheet.cell(row=totalDesa, column=3).value = data_e_hora_em_texto

            wb.save(str(file))

        total = total + 1
        print('A última ação verificada foi a ', i)
        print('O total de ações verificada é de: ', total)

    print('Lista de ações com datas atualizadas possuem um total de: ', len(listAtuais))
    print('Lista de ações com datas desatualizadas possuem um total de: ', len(listAntigo))
