# Básicamente precisaremos do script e de uma planilha xlsx para deixarmos nossos dados organizados
# Para o adequado funcionamento deste código é necessária a instalação dos pacotes (ou modulos?) abaixo
import shutil
import os

import openpyxl

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

from datetime import datetime
import time

# Antes vamos organizar as variáveis de entrada, indicando:
# 1 - o diretório onde gostariamos de salvar nosso
# 2 - Uma data para que nosso código verifique se os dados são do ultimo dia de cotação
# 3 - O nome de nossa arquivo excel que receberá os dados
# 4 - A planilha do arquivo onde inseriremos os dados atualizados
# 5 - A planilha do arquivo onde inseriremos os dados desatualizados

path = "C:\\Users\\Usuário\\Desktop"
diaNegocios = "13/11/2020"
file = 'diariamente.xlsx'
planilha01 = 'Atualizadas'
planilha02 = 'Desatualizadas'

# A seguir 3 linhas de código para que seja feita uma cópia do arquivo e a mesma copia seja colada na área de trabalho
original = r'' + str(file)
alvo = r'' + str(path) + '\\' + str(file)
shutil.copyfile(original, alvo)

# Agora vamos pedir para que o nosso código trabalhe apenas na área de trabalho
os.chdir(path)

# Por ultimo o código reedita a string diaNegocios retirando as barras inclinadas
# Fazemos isso para que o arquivo excel seja renomeado com a data que desejamos dos dados
# Essa edição é gerada uma variável como ultimoDia
# os.rename, completa a mudança do nome do arquivo

ultimoDia = diaNegocios.replace('/', '_')
os.rename(r'' + str(file), r'' + str(ultimoDia) + '.xlsx')

# A ultima edição antes da soleta de dados é da variável file, para que o arquivo seja reconhecido no restante do código

file = str(ultimoDia) + '.xlsx'

# A seguir são criadas três listas vazias que usaremos mais adiante para armazenamento e monitoramentos, são elas:
# listatuais - armazenar o nome das ações que estão com dados atualizados do ultimo dia de negociações
# listAntigo - armazenar o nome das ações que não estão com dados atualizados do ultimo dia de negociações
# listStocks - armazenará a lista de ações listada na planilha diariamente.xlsx

listAtuais = []
listAntigo = []
listStocks = []

# A seguir iniciamos o trabalho de acesso ao arquivo .xlsx:
# O modulo load_workbook() da função openpyxl abre o arquivo .xlsx
# A planilha "Atualizadas" é acessada pelo wb[str(planilha01)] e armazenada na variável sheet
# A partir de então o arquivo estará sendo editado

wb = openpyxl.load_workbook(str(file))
sheet = wb[str(planilha01)]

# O laço for irá trabalhar nela
# 1- Dentro do laço o i será substituido com do número 2 até o 350
# 2- Desta forma sheet.cell irá acessar linha por linha, da 2 a 322, e em cada linha capiturará o valor dela e
# aramazenará na variável "nameStock"
# 3- o modulo append() colocará cada valor assumido por nameStock, a cada loop for, na lista listStocks,
# que foi inicalmente criada como uma lista vazia
# ao final teremos a lista "listStocks" com todas ações armazenada nela

for i in range(2, 400, 1):
    nameStock = sheet.cell(row=i, column=1).value
    listStocks.append(nameStock)

# Abaixo criamos dois contadores que seram usado nos código de capitura e monitoramento

total = 0  # Representa o total de ações que foram verificadas no site fundamentus a cada looping

j = 1  # Representa o total de ações verificadas no site fundamentus, com datas do ultimo dia de negociações e
# listadas na planilha "Ataulizadas"

totalDesa = 1  # Representa o total de ações verificadas no site fundamentus, com data desatualizada e listadas na
# planilha "desatuzalidaa"

# As linhas a seguir
# 1 - Amazena a função que abre o Chrome, na variável fund
# 2 - Chama o modulo get() para abrirá o navegador e acessar o endereço indicado
# 3 - Chama a função maximize_window() para maximizar a tela

fund = webdriver.Chrome(ChromeDriverManager().install())
fund.get("http://www.fundamentus.com.br/detalhes.php?papel=")
fund.maximize_window()

# No laço for a seguir:
# Irá passar por cada um das siglas da ações inicialmente armazenadas na listStocks
# find_element_by_id('completar') encontra a caixa no site fundamentus para inserir a sigla das ações e armazamos
# a localização dessa caixa na variável "caixa"
# send_keys(i) escreve o nome da açao i na "caixa"
# login_attempt encontra o botão para acessar os dados da ação de interesse
# o botão é clicado ao ser aplicado o modulo .submit()
# time.sleep(1.5) é um intervalo inserido para que dê tempo de um completo carregamento da
# página onde se encontra os dados
for i in listStocks:
    caixa = fund.find_element_by_id('completar')
    caixa.send_keys(i)
    login_attempt = fund.find_element_by_xpath('/html/body/div[1]/div[1]/form/fieldset/input[2]')
    login_attempt.submit()
    time.sleep(1.5)

    # Dado que algumas das ações listadas podem não apresentar dados e consequentemente não abrir a página de interesse,
    # conduzindo a erros do nosso processamento do código então nos inserimos um "try" e dentro dele tentamos encontrar
    # o dado básico que é a data dos dados
    # Se a data for encontrada então a capitura de dados segue normalmente
    # Do contrário o código passa para a próxima açõa de nossa lista...
    try:
        data = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[2]/td[4]').text
        

    # Ou seja, Caso o erro ocorra então o "except" imprimirá no console o código abaixo e o código passa para a próxima
    # açõa de nossa lista
    except:
        print('Opa tivemos um erro aqui com a ação ' + i)

    # O else execultará o código caso a data seja encontrada
    else:
        # Condição "if":
        # Se a data for encontrada com o valor determinado las nas primeiras linhsa do nosso código, na variável
        # diaNegocios então a ação será armazenada na lista "listAtuais"
        # e a condição "if" continuará sendo execultada capiturando os dados que desejamos
        if data == diaNegocios:
            listAtuais.append(i)

            # Os códigos abaixo, de capitura de dados, pode realizar dois tipos de tratamento:
            # Primeira possibilidade de tranformação:
            # 1 - usando ".replace" para substituir "," (virgulas) por "." (pontos)

            # Segunda possibilidade de tranformação:
            # 2 - um laço "for" que também faz o mesmo tipo de substituição, contudo ocorre da seguinte forma:
            # 2.1 - Tranformamos o conteúdo da variavem em uma lista
            # 2.1 - O laço então primeiro elimina todos os pontos, inclusive dos agrupadores de digitos
            # 2.2 - depois ele substitui a virgula "," por ".", neste caso do separador decimal
            # Obs: ambos tipos de substituição são realizadas para que o excel reconheça corretamente os dados inseridos
            # neles já que no meu sistema operacional as configuração estão para reconhecer "." como separadore decimal
            # Esta segunda forma de transformação serve para número decimais que possui mais que 999,00

            # Variação dia
            vard = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[2]/td[2]').text
            vard_n = vard.replace(',', '.')

            # Variação mês
            varm = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[3]/td[2]/span/font').text
            varm_n = varm.replace(',', '.')

            # Variação 30 dias
            var30d = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[4]/td[2]/span/font').text
            var30d_n = var30d.replace(',', '.')

            # Variação 12 meses
            var12m = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[5]/td[2]/span/font').text
            var12m_n = list(var12m)
            for x, w in zip(var12m_n, range(0, len(var12m_n), 1)):
                if x == '.':
                    del var12m_n[w]
                elif x == ',':
                    var12m_n[w] = '.'
            var12m_n = ''.join(var12m_n)
            var12m_n = str(var12m_n)

            # Variação no ano de 2020
            var2020 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[6]/td[2]/span/font').text
            var2020_n = list(var2020)
            for x, w in zip(var2020_n, range(0, len(var2020_n), 1)):
                if x == '.':
                    del var2020_n[w]
                elif x == ',':
                    var2020_n[w] = '.'
            var2020_n = ''.join(var2020_n)
            var2020_n = str(var2020_n)

            # Variação no ano de 2019
            var2019 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[7]/td[2]/span/font').text
            var2019_n = list(var2019)
            for x, w in zip(var2019_n, range(0, len(var2019_n), 1)):
                if x == '.':
                    del var2019_n[w]
                elif x == ',':
                    var2019_n[w] = '.'
            var2019_n = ''.join(var2019_n)
            var2019_n = str(var2019_n)

            # Variação no ano de 2018
            var2018 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[8]/td[2]/span/font').text
            var2018_n = list(var2018)
            for x, w in zip(var2018_n, range(0, len(var2018_n), 1)):
                if x == '.':
                    del var2018_n[w]
                elif x == ',':
                    var2018_n[w] = '.'
            var2018_n = ''.join(var2018_n)
            var2018_n = str(var2018_n)

            # Variação no ano de 2017
            var2017 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[9]/td[2]/span/font').text
            var2017_n = list(var2017)
            for x, w in zip(var2017_n, range(0, len(var2017_n), 1)):
                if x == '.':
                    del var2017_n[w]
                elif x == ',':
                    var2017_n[w] = '.'
            var2017_n = ''.join(var2017_n)
            var2017_n = str(var2017_n)

            # Variação no ano de 2016
            var2016 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[10]/td[2]/span/font').text
            var2016_n = list(var2016)
            for x, w in zip(var2016_n, range(0, len(var2016_n), 1)):
                if x == '.':
                    del var2016_n[w]
                elif x == ',':
                    var2016_n[w] = '.'
            var2016_n = ''.join(var2016_n)
            var2016_n = str(var2016_n)

            # Variação no ano de 2015
            var2015 = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[11]/td[2]/span/font').text
            var2015_n = list(var2015)
            for x, w in zip(var2015_n, range(0, len(var2015_n), 1)):
                if x == '.':
                    del var2015_n[w]
                elif x == ',':
                    var2015_n[w] = '.'
            var2015_n = ''.join(var2015_n)
            var2015_n = str(var2015_n)

            # VPA
            vpa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[3]/td[6]/span').text
            vpa_n = vpa.replace(',', '.')

            # ROE
            roe = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[9]/td[6]/span').text
            roe_n = roe.replace(',', '.')

            # PA
            pa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[1]/td[4]').text
            pa_n = pa.replace(',', '.')

            # P/VPA
            p_vpa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[3]/td[4]').text
            p_vpa_n = p_vpa.replace(',', '.')

            # PL
            pl = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[2]/td[4]').text
            pl_n = pl.replace(',', '.')

            # LPA
            lpa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[2]/td[6]').text
            lpa_n = lpa.replace(',', '.')

            # DIV. BRT/PATR
            divbppa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[11]/td[6]').text
            divbppa_n = divbppa.replace(',', '.')

            # P/EBIT
            p_ebit = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[4]/td[4]').text
            p_ebit_n = p_ebit.replace(',', '.')

            # Div Yield
            div_yield = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[3]/tbody/tr[9]/td[4]/span').text
            div_yield_n = div_yield.replace(',', '.')

            # Setor da Empresa
            setor = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[4]/td[2]').text

            # Nome da Empresa
            empresa = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[3]/td[2]').text

            # Data do ultimo balanço
            balanco = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[2]/tbody/tr[1]/td[4]').text

            # A seguir:
            # A planilha "Atualizadas" é acessada pelo wb[str(planilha01)] e armazenada na variável sheet
            sheet = wb[str(planilha01)]

            # Os dados são então armazenados numa lista chamada "listDados"
            listDados = [i, empresa, setor, roe_n, vpa_n, pa_n, p_vpa_n, pl_n, lpa_n,
                         divbppa_n, p_ebit_n, div_yield_n, vard_n, varm_n, var30d_n,
                         var12m_n, var2020_n, var2019_n,
                         var2018_n, var2017_n, var2016_n, var2015_n, balanco, data]

            # Antes de entrar no "for" que irá levar os dados para planilha o contador J recebe mais uma unidade
            # Para que ele começe na linha correspondente à ação trabalhando em cada looping
            j = j + 1

            # No laço "for":
            # Cada indice da "listDados" será acessado através do indice "coluna"
            # j representa a linha que receberá os dados durante o "for"
            # O indice coluna fala duas funções
            # 1 - A cada looping ele indicará em qual coluna deve ser impressa a informção de lista de dados
            # 2 - A cada looping ele imprimirá um dado diferente de cada um dos itens da lista dados
            # Por ultimo, para auxiliar no monitoramento usamos as funções de "datatime" e armazenamos o horário de
            # coleta na ultima coluna de dados
            for coluna in range(0, len(listDados), 1):
                sheet.cell(row=j, column=coluna + 1).value = listDados[coluna]
                data_e_hora_atuais = datetime.now()
                data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')
                sheet.cell(row=j, column=len(listDados) + 1).value = data_e_hora_em_texto

            # O arquivo é salvo assim que completada a condição elif
            wb.save(str(file))

        # Contudo caso, no inicio do nossa condição, if,a data seja diferente da data determinada
        # a ação i é armazanada na lista "listAntigo", através do elif abaixo

        elif data != 'diaNegocios':
            listAntigo.append(i)

            # A seguir:
            # A planilha "Atualizadas" é acessada pelo wb[str(planilha02)] e armazenada na variável sheet
            sheet = wb[str(planilha02)]

            # Os dados são então armazenados numa lista chamada "listDados"
            listDados = [i, data]

            # Antes de entrar no "for" que irá levar os dados para planilha o contador totalDesa recebe mais uma unidade
            # Para que ele começe numa linha diferente para cada à ação trabalhada em cada looping
            totalDesa = totalDesa + 1

            # No laço "for":
            # Cada indice da "listDados" será acessado através do indice "coluna"
            # totalDesa representa a linha que receberá os dados durante o "for"
            # O indice coluna fala duas funções
            # 1 - A cada looping ele indicará em qual coluna deve ser impressa a informção de lista de dados
            # 2 - A cada looping ele imprimirá um dado diferente de cada um dos itens da lista dados
            # Por ultimo, para auxiliar no monitoramento usamos as funções de "datatime" e armazenamos o horário de
            # coleta na ultima coluna de dados
            for coluna in range(0, len(listDados), 1):
                sheet.cell(row=totalDesa, column=coluna + 1).value = listDados[coluna]
                data_e_hora_atuais = datetime.now()
                data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')
                sheet.cell(row=totalDesa, column=3).value = data_e_hora_em_texto

            # O arquivo é salvo assim que completada a condição elif
            wb.save(str(file))

        # Para fins de monitoramento as linhas abaixo foram criadas nas respectivas identações
        # A cada ação verificada, independente do dado ser atualizado ou não, é somado +1 no contador "conta
        total = total + 1
        print('A ultima ação verificada foi a ', i)
        print('O total de ações verificada é de: ', total)

    print('Lista de ações com datas atualizadas possuem um total de: ', len(listAtuais))
    print('Lista de ações com datas desatualizadas possuem um total de: ', len(listAntigo))
