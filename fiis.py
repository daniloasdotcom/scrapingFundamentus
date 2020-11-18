import pprint
import os
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
from gtts import gTTS
from playsound import playsound
import shutil
import send2trash

path = "C:\\Users\\Usuário\\Desktop"

file = 'fiis.xlsx'
planilha01 = 'Planilha1'
listaFiis = []

original = r'' + str(file)
alvo = r'' + str(path) + '\\' + str(file)

shutil.copyfile(original, alvo)

os.chdir(path)

wb = openpyxl.load_workbook(str(file))
sheet = wb[str(planilha01)]

for i in range(2, 244, 1):
    nameStock = sheet.cell(row=i, column=1).value
    listaFiis.append(nameStock)

pprint.pprint(listaFiis)

page = webdriver.Chrome(ChromeDriverManager().install())

listaErros = []
dic = {}
x = 1
for i in listaFiis:
    i = str(i)
    i = i.lower()
    i = i.replace('\n', '')
    path = 'https://fiis.com.br/' + i
    page.get(path)
    time.sleep(5)
    x = x + 1
    print('Iniciando a abertura dos dados do fundo:', i.upper())
    print('Através do endereço: ', path)
    try:
        basic = '//*[@id="quotations--infos-wrapper"]/div[1]/span[2]'
        teste = page.find_element_by_xpath(basic).text

    except:
        print('############################')
        print('ATENÇÂO!   ATENÇÂO!   ATENÇÂO!')
        print('############################')
        print('Tivemos um erro com o fundo:', i.upper())
        print('O endereço de acesso é', path)
        print(' ')
        print(' ')
        listaErros.append(path)
        dic.update({str(path): x})

    else:
        print('Capiturando dados do fundo...')

        # Cotação Atual
        basic = '//*[@id="quotations--infos-wrapper"]/div[1]/span[2]'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=x, column=13).value = download

        wb.save(str(file))

        # Variação
        # basic = '//*[@id="quotations--infos-wrapper"]/div[1]/div'
        # download = page.find_element_by_xpath(basic).text
        # sheet.cell(row=k, column=14).value = download

        # wb.save(str(file))

        # endereço dos dados
        sheet.cell(row=x, column=15).value = path

        wb.save(str(file))

        # Administrador
        basic = '//*[@id="informations--admin"]/div[1]/div[2]/span[1]'

        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=x, column=16).value = download

        wb.save(str(file))

        # Telefone
        basic = '//*[@id="informations--admin"]/div[2]/div[1]/div[2]/span[2]'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=x, column=17).value = download

        wb.save(str(file))

        # e-mail
        basic = '//*[@id="head--card"]/div[2]/div[2]/div[2]/div[2]/span[2]/a'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=x, column=18).value = download

        wb.save(str(file))

        # Site
        basic = '//*[@id="head--card"]/div[2]/div[2]/div[3]/div[2]/span[2]/a'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=x, column=19).value = download

        wb.save(str(file))

        # Nome no pregão, Tipo do FII, Tipo AMBIMA, REGSITRO CVM
        for j, z in zip(range(1, 5, 1), range(2, 6, 1)):
            basic = '//*[@id="informations--basic"]/div[1]/div[' + str(j) + ']/span[2]'
            download = page.find_element_by_xpath(basic).text
            sheet.cell(row=x, column=z).value = download
        wb.save(str(file))

        # Nº de Cotas, Número de cotistas, CNPJ
        for j, z in zip(range(1, 4, 1), range(6, 9, 1)):
            basic = '//*[@id="informations--basic"]/div[2]/div[' + str(j) + ']/span[2]'
            download = page.find_element_by_xpath(basic).text
            sheet.cell(row=x, column=z).value = download
        wb.save(str(file))

        #
        for j, z in zip(range(1, 5, 1), range(9, 14, 1)):
            basic = '//*[@id="informations--indexes"]/div[' + str(j) + ']/span[1]'
            download = page.find_element_by_xpath(basic).text
            sheet.cell(row=x, column=z).value = download
        wb.save(str(file))
        print('Dados capiturados e aparentemente tudo ok')
        print('Total de fundos varificados:', (x - 1))
        print(' ')
        print('Endereços com problema de acesso aos dados: ', )
        pprint.pprint(listaErros)
        print(' ')
        print('Endereço / Linha para inserir os dados', )
        pprint.pprint(dic)
        print(' ')


print('A seguir iniciaremos uma novo acesso às páginas que apresentaram erro:')
print('############################')
print('############################')
print('')
print('')

listaErros02 = []
dic02 = {}

# tratando erros
for v, k in dic.items():
    print('Vamos testar o acesso à página: ', v)
    try:
        page.get(v)
        time.sleep(5)
        basic = '//*[@id="quotations--infos-wrapper"]/div[1]/span[2]'
        teste = page.find_element_by_xpath(basic).text

    except:
        print('Continuamos tendo um erro com o fundo: ', v)
        print('O endereço de acesso é', path)
        listaErros02.append(v)
        dic02.update({str(path): x})

    else:
        print('O endereço abaixo não está com problemas:')
        print(v)
        print('Capiturando os dados...')

        # Cotação Atual
        basic = '//*[@id="quotations--infos-wrapper"]/div[1]/span[2]'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=k, column=13).value = download

        wb.save(str(file))

        # Variação
        # basic = '//*[@id="quotations--infos-wrapper"]/div[1]/div'
        # download = page.find_element_by_xpath(basic).text
        # sheet.cell(row=k, column=14).value = download

        # wb.save(str(file))

        # endereço dos dados
        sheet.cell(row=k, column=15).value = v

        wb.save(str(file))

        # Administrador
        basic = '//*[@id="head--card"]/div[2]/div[1]/div[2]/span[2]'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=k, column=16).value = download

        wb.save(str(file))

        # Telegone
        basic = '//*[@id="head--card"]/div[2]/div[2]/div[1]/div[2]/span[2]'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=k, column=17).value = download

        wb.save(str(file))

        # e-mail
        basic = '//*[@id="head--card"]/div[2]/div[2]/div[2]/div[2]/span[2]/a'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=k, column=18).value = download

        wb.save(str(file))

        # Site
        basic = '//*[@id="head--card"]/div[2]/div[2]/div[3]/div[2]/span[2]/a'
        download = page.find_element_by_xpath(basic).text
        sheet.cell(row=k, column=19).value = download

        wb.save(str(file))

        # Nome no pregão, Tipo do FII, Tipo AMBIMA, REGSITRO CVM
        for j, z in zip(range(1, 5, 1), range(2, 6, 1)):
            basic = '//*[@id="informations--basic"]/div[1]/div[' + str(j) + ']/span[2]'
            download = page.find_element_by_xpath(basic).text
            sheet.cell(row=k, column=z).value = download
        wb.save(str(file))

        # Nº de Cotas, Número de cotistas, CNPJ
        for j, z in zip(range(1, 4, 1), range(6, 9, 1)):
            basic = '//*[@id="informations--basic"]/div[2]/div[' + str(j) + ']/span[2]'
            download = page.find_element_by_xpath(basic).text
            sheet.cell(row=k, column=z).value = download
        wb.save(str(file))

        # Dividend Yield, Último Rendimento, Patrimônio Liquido, Valor patrimonial/Cota
        for j, z in zip(range(1, 5, 1), range(9, 14, 1)):
            basic = '//*[@id="informations--indexes"]/div[' + str(j) + ']/span[1]'
            download = page.find_element_by_xpath(basic).text
            sheet.cell(row=k, column=z).value = download
        wb.save(str(file))
        print('Dados capiturados e aparentemente tudo ok')
        print(' ')
        print('Relatório de erros')
        print('Endereços com problema de acesso aos dados: ', )
        pprint.pprint(listaErros02)
        print(' ')
        print('Endereço / Linha para inserir os dados', )
        pprint.pprint(dic02)
        print(' ')

print('Favor, rever os erros')
print('Tente verficar se os endereços listados estão com os digitos corretos')
num_de_cell = str('Danilo, Finalizei a coleta dos dados de fundo imobiliários')
tts = gTTS(num_de_cell, lang='pt-br')
tts.save('num_de_cell.mp3')  # Salva o arquivo de audio
playsound('num_de_cell.mp3')
send2trash.send2trash('nome_arquivo.mp3')