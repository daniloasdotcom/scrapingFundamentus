#Script para acessar e verificar toda a lista de ações do site Fundamentus

import openpyxl
import os
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
import time

path = "C:\\Users\\Usuário\\Desktop"
os.chdir(path)

fund = webdriver.Chrome(ChromeDriverManager().install())
fund.get("http://www.fundamentus.com.br/detalhes.php?papel=")
fund.maximize_window()

j = 1 # -Contador de ações atuais e indicador de linha para impressão do nome e data da ação
h = 1 # -Contador de ações antigas e indicador de linha para impressão do nome e data da ação
y = 1 # -Contador de ações sem informações e indicador de linha para impressão do nome e data da ação

listStocks = []
listAntigo = []
listAtuais = []

for i in range(h,970,1):
    i = str(i)
    path = '//*[@id="test1"]/tbody/tr[' + i + ']/td[1]/a'
    nameStock = fund.find_element_by_xpath(path).text
    listStocks.append(nameStock)

for i in listStocks:
        comentar = fund.find_element_by_id('completar')
        comentar.send_keys(i)
        login_attempt = fund.find_element_by_xpath('/html/body/div[1]/div[1]/form/fieldset/input[2]')
        login_attempt.submit()
        time.sleep(1.5)

        try:
            data = fund.find_element_by_xpath('/html/body/div[1]/div[2]/table[1]/tbody/tr[2]/td[4]').text
        except:
            print('Opa tivemos um erro aqui com a ação ' + i)
            y = y + 1
            wb = openpyxl.load_workbook('stocks.xlsx')
            sheet = wb.get_sheet_by_name('Planilha1')
            sheet.cell(row=y, column=5).value = i
            wb.save('stocks.xlsx')

            print(' Total de ações sem informação impressas: ' + str(y - 1))
            print(' Total de ações antigas impressas: ' + str(h - 1))
            print(' Total de ações atuais impressas: ' + str(j - 1))
            total = (h - 1) + (y - 1) + (j - 1)
            print(' Total de ações averiguadas: ' + str(total))

        else:
            if data == '01/03/2021':
                listAtuais.append(i)
                j = j + 1

                print(listAtuais)

                wb = openpyxl.load_workbook('stocks.xlsx')
                sheet = wb.get_sheet_by_name('Planilha1')
                sheet.cell(row=j, column=1).value = i
                sheet.cell(row=j, column=2).value = data
                wb.save('stocks.xlsx')

                print(' Total de ações sem informação impressas: ' + str(y - 1))
                print(' Total de ações antigas impressas: ' + str(h - 1))
                print(' Total de ações atuais impressas: ' + str(j - 1))
                total = (h - 1) + (y - 1) + (j - 1)
                print(' Total de ações averiguadas: ' + str(total))

            elif data != '01/03/2021':
                listAntigo.append(i)
                h = h + 1
                wb = openpyxl.load_workbook('stocks.xlsx')
                sheet = wb.get_sheet_by_name('Planilha1')
                sheet.cell(row=h, column=3).value = i
                sheet.cell(row=h, column=4).value = data

                wb.save('stocks.xlsx')

                print('Ação: ' + i + ' de data igual à ' + data)

                print(' Total de ações sem informação impressas: ' + str(y-1))
                print(' Total de ações antigas impressas: ' + str(h-1))
                print(' Total de ações atuais impressas: ' + str(j-1))
                total = (h - 1) + (y - 1) + (j - 1)
                print(' Total de ações averiguadas: ' + str(total))
